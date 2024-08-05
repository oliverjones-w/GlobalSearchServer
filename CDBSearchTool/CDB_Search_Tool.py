

#When a name is fed into this script it will create two dictionaries; one with CDB values, and one with PDB values
#This script needs to be modified so that it creates a dictionary for all the duplicate values as well..... but for now, that's okay

import argparse
import openpyxl
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

#define the centralized database
centralizedDB = r"K:\Market Maps\Centralized Database.xlsx"

#Set the paths for the various maps
CreditPath = r"K:\Market Maps\Credit Map.xlsm"
RatesPath = r"K:\Market Maps\Interest Rates Map.xlsm"
CommoditiesPath = r"K:\Market Maps\Commodites Map.xlsm"
HedgeFundPath = r"C:\Users\BSA-OliverJ'22\OneDrive\Desktop\OneDrive\Mapping\Hedge Fund Map (Personal).xlsm"
EquitiesPath = r"K:\Market Maps\Equities Map.xlsx"

#Instantiate a dictionary linking the keywords to the appropriate path

Map_Path_Dict = {'Credit': CreditPath, 
                 'Interest Rates': RatesPath, 
                 'Commodities': CommoditiesPath,
                 'Hedge Fund': HedgeFundPath, 
                 'Equities': EquitiesPath
                }

#Define the name to search using the user's input
parser = argparse.ArgumentParser(description="Search for a name in the database")
parser.add_argument("name_to_search", type=str, help="Name to search in the database")
args = parser.parse_args()

name_to_search = args.name_to_search

#This function creates a dictionary linking headers to header indices for a given database

def get_header_indices(wb_path, sheet_name, max_header_row):
    try: 
        wb = openpyxl.load_workbook(wb_path)
        ws = wb[sheet_name]

        header_indices = {}  # Dictionary to store headers and indices
        header_row_found = False

        for row in ws.iter_rows(max_row=max_header_row):
            # Check if this row contains the "Name" cell
            if any(cell.value == "Name" for cell in row):
                header_row_found = True
                for col_idx, cell in enumerate(row, start=1):
                    header_name = str(cell.value)
                    if header_name:
                        header_indices[header_name] = col_idx
                break  # Exit the loop once the header row is processed

        if not header_indices:
            raise ValueError("No headers were found in the first {} rows of the '{}' sheet.".format(max_header_row, sheet_name))
        if not header_row_found:
            raise ValueError("The header row containing 'Name' was not found in the first {} rows.".format(max_header_row))

        return header_indices

    except FileNotFoundError:
        raise FileNotFoundError("The file '{}' was not found.".format(wb_path))
    except KeyError:
        raise KeyError("The sheet '{}' was not found in the workbook.".format(sheet_name))

#This function creates a dictionary based on the location of the input name

def search_name_in_database(name_to_search, header_indices, wb_path, sheet_name):
    try:
        # Get the column index for the 'Name' header from the provided dictionary
        name_column_index = header_indices.get("Name")

        if name_column_index is not None:
            # Open the workbook
            wb = openpyxl.load_workbook(wb_path, data_only=True)
            ws = wb[sheet_name]

            # Initialize a variable to keep track of whether the name was found
            name_found = False
            # Initialize the results dictionary
            results = {}

            # Iterate through rows to find the name
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[name_column_index - 1] and row[name_column_index - 1].lower() == name_to_search.lower():
                    name_found = True
                    # Populate the results dictionary
                    for key in header_indices:
                        results[key] = row[header_indices[key] - 1]
                    break  # Stop searching after finding the name

            if not name_found:
                results["error"] = f"The name '{name_to_search}' was not found in the database"
            return results

        else:
            return {"error": "The 'Name' column index was not found in the provided header dictionary"}

    except Exception as e:
        return {"error": str(e)}

#Create the CDB dictionary
CDBheader_indices = get_header_indices(centralizedDB, "Sheet", 5)
CDB_results = search_name_in_database(name_to_search, CDBheader_indices, centralizedDB, "Sheet")

#Create the PDB dictionary
PDB_header_indices = get_header_indices(Map_Path_Dict[CDB_results["Source"]], "Master", 10)
PDB_results = search_name_in_database(name_to_search, PDB_header_indices, Map_Path_Dict[CDB_results["Source"]], "Master")

print(CDB_results["Firm"])
print(CDB_results["Title"])
print("Map: ", CDB_results["Source"])
print(PDB_results["Location"])
if PDB_results["Code"] is not None and PDB_results["Mobile"] is not None:
    code = str(PDB_results["Code"]).strip()
    mobile = str(PDB_results["Mobile"]).strip()
    print(f"Mobile: {code} {mobile}")